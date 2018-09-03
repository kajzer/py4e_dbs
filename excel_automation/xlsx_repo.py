#!/usr/bin/env python3
# xlsx_repo.py - showcase of different functions

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

def main():
    wb = openpyxl.load_workbook('Book1.xlsx')
    # Get sheet names as a list
    print(wb.sheetnames)
    
    # get sheet object
    sheet = wb['Sheet1']
    print(sheet)
    print(type(sheet))
    
    # get sheet title
    print(sheet.title)
    
    # get active sheet
    activeSheet = wb.active
    print(activeSheet)
    print('====Getting Cells From The Sheet=====')
    
    # get cell object
    print(sheet['A1'])
    
    # get value
    print(sheet['A1'].value)
    c = sheet['B1']
    print(c.value)
    print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
    print('Cell ' + c.coordinate + ' is ' + c.value)
    
    # navigating sheet
    print('====Navigating======')
    print(sheet.cell(row=1, column=2).value)
    
    for i in range(1, 8, 2):
        print(i, sheet.cell(row=i, column=2).value)
        
    # determine sheet size
    print(sheet.max_row, sheet.max_column)
    
    # converting between column letters and numbers
    print(get_column_letter(1))
    print(get_column_letter(27))
    print(get_column_letter(sheet.max_column))
    print(column_index_from_string('AA'))
    
    # getting rows and columns from the sheet
    print(tuple(sheet['A1:C3']))
    
    # display contents in range
    for rowOfCellObjects in sheet['A1:C3']:
        for cellObj in rowOfCellObjects:
            print(cellObj.coordinate, cellObj.value)
        print('--- END ROW ---')
        
    for cellObj in sheet.columns[1]:
        print(cellObj.value)
    
if __name__ == "__main__":
    main()